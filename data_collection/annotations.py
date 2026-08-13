#!/usr/bin/env python3
"""
Per-episode annotation store + the pygame prompt that fills it.

Everything a human knows about a run — did it work, was the grape damaged,
was anything odd — had nowhere to go before this. exclude.txt carries exactly
two facts (drop, crop) and is written by the curation scripts, long after the
person who watched the run has forgotten what they saw.

This records the perishable half at the keyboard, the moment the episode is
saved, and writes it next to the episodes it describes:

    demo_data/annotations.csv     append-only, stdlib csv, the source of truth
    demo_data/annotations.xlsx    regenerated after every row, for reading

WHY BOTH. The csv is written with the standard library, so a missing package
can never cost you an annotation, and an append is a single line — a crash
mid-write loses at most the row in flight. The xlsx is a convenience view
regenerated from it; if openpyxl is not installed, or the workbook is open in
Excel and locked, the write is skipped with a warning and the csv is still
complete. Never treat the xlsx as the record: edit the csv, or re-export.

    /usr/bin/python3.12 -m pip install --break-system-packages openpyxl

WHAT GOES IN THE PROMPT. Only what the footage cannot recover later. The vine
hung low, the operator's hand slipped, tracking glitched mid-reach — those
exist for about ten seconds and then they are gone. Grading that needs a close
look at the grape belongs here too *while the operator is holding it*, but
anything needing the recording reviewed belongs in a later review pass, not at
the keyboard with a live arm two metres away. Fields are declared in
tasks/<task>.yaml so adding one is a config change.

ESC ALWAYS WORKS. Skipping writes status=unreviewed rather than nothing, so an
un-annotated episode is a visible state you can filter for later, not a silent
gap. R / S / Q also escape the prompt (saving whatever was answered) so the
prompt can never stand between the operator and the robot controls.
"""

import csv
import os
import subprocess
from datetime import datetime

import yaml

# Keys the collector owns. A task spec may not bind these as answer hotkeys or
# the operator loses the ability to reset/start/quit from inside the prompt.
RESERVED_KEYS = set('rsdcpuq')
TEXT_KEY = 't'

TASKS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'tasks')

# Written before the task's own fields, then after them. Keeping provenance in
# the same row as the judgement is the point: "were the runs recorded under the
# -48 degree pitch clamp worse?" is a question you cannot ask retrospectively
# unless the setting was written down at the time.
CORE_HEAD = ['episode', 'recorded_at', 'task', 'operator', 'status']
CORE_TAIL = ['notes', 'num_frames', 'duration_s', 'rate_hz',
             'sensors', 'git_sha', 'git_dirty']

STATUS_ANNOTATED = 'keep'
STATUS_SKIPPED = 'unreviewed'
STATUS_DELETED = 'deleted'        # row kept, episode_N.hdf5 no longer on disk
STATUS_SUPERSEDED = 'superseded'  # an earlier row for a reused episode number


# ── Task spec ──────────────────────────────────────────────────────────────────

class Field:
    """One question. `choices` is a list of (value, key) pairs for enums."""

    def __init__(self, d: dict):
        self.name = d['field']
        self.ask = d.get('ask', self.name)
        self.type = d.get('type', 'enum')
        self.rule = d.get('rule', '')
        self.choices = []
        for c in d.get('choices', []):
            if isinstance(c, dict):
                self.choices.append((str(c['value']), str(c['key']).lower()))
            else:
                self.choices.append((str(c), str(c)[0].lower()))

    @property
    def keys(self):
        return [k for _, k in self.choices]

    def value_for_key(self, key):
        for value, k in self.choices:
            if k == key:
                return value
        return None


class TaskSpec:
    """Loaded from tasks/<name>.yaml. Only the fields marked `when: collect`
    (the default here) are asked at the keyboard."""

    def __init__(self, name, display, fields):
        self.name = name
        self.display = display
        self.fields = fields

    @classmethod
    def load(cls, name_or_path: str):
        path = name_or_path
        if not os.path.isfile(path):
            path = os.path.join(TASKS_DIR, f'{name_or_path}.yaml')
        with open(path) as fh:
            spec = yaml.safe_load(fh) or {}

        fields, seen = [], {}
        for raw in spec.get('annotations', []):
            if raw.get('when', 'collect') != 'collect':
                continue                     # graded later, not at the keyboard
            f = Field(raw)
            for k in f.keys:
                if k in RESERVED_KEYS or k == TEXT_KEY:
                    raise ValueError(
                        f"{path}: field '{f.name}' binds reserved key '{k}' "
                        f"(reserved: {sorted(RESERVED_KEYS)} + '{TEXT_KEY}')")
                if k in seen:
                    raise ValueError(
                        f"{path}: key '{k}' bound by both '{seen[k]}' and '{f.name}'")
                seen[k] = f.name
            fields.append(f)

        return cls(spec.get('name', name_or_path),
                   spec.get('display', spec.get('name', name_or_path)),
                   fields)


# ── Store ──────────────────────────────────────────────────────────────────────

def _git_state(repo_dir):
    """(short sha, dirty flag) for the code that recorded this episode."""
    def run(*args):
        return subprocess.run(args, cwd=repo_dir, capture_output=True,
                              text=True, timeout=5).stdout.strip()
    try:
        sha = run('git', 'rev-parse', '--short', 'HEAD')
        dirty = 'yes' if run('git', 'status', '--porcelain') else 'no'
        return sha, dirty
    except Exception:
        return '', ''


class AnnotationStore:
    """One csv (+ mirrored xlsx) per collection folder, alongside the episodes."""

    def __init__(self, save_dir: str, spec: TaskSpec, log=None):
        self._dir = save_dir
        self._spec = spec
        self._log = log
        self._csv = os.path.join(save_dir, 'annotations.csv')
        self._xlsx = os.path.join(save_dir, 'annotations.xlsx')
        self._xlsx_warned = False
        self.columns = CORE_HEAD + [f.name for f in spec.fields] + CORE_TAIL
        self.git_sha, self.git_dirty = _git_state(
            os.path.dirname(os.path.abspath(__file__)))

    def _info(self, msg):
        (self._log.info if self._log else print)(msg)

    def _warn(self, msg):
        (self._log.warn if self._log else print)(msg)

    def annotated_episodes(self) -> set:
        """Episode names already in the csv — so a restart resumes rather than
        duplicating, and a later review pass can find the gaps."""
        if not os.path.isfile(self._csv):
            return set()
        try:
            with open(self._csv, newline='') as fh:
                return {r['episode'] for r in csv.DictReader(fh) if r.get('episode')}
        except Exception:
            return set()

    def read_rows(self) -> list:
        if not os.path.isfile(self._csv):
            return []
        with open(self._csv, newline='') as fh:
            return list(csv.DictReader(fh))

    def _rewrite(self, rows: list):
        """Rewrite the whole csv atomically. Only for status corrections —
        normal recording appends."""
        tmp = self._csv + '.tmp'
        with open(tmp, 'w', newline='') as fh:
            w = csv.DictWriter(fh, fieldnames=self.columns, extrasaction='ignore')
            w.writeheader()
            for r in rows:
                w.writerow({c: r.get(c, '') for c in self.columns})
        os.replace(tmp, self._csv)

    def orphan_rows(self) -> list:
        """Rows whose episode_N.hdf5 is no longer on disk.

        Deleting a bad episode by hand is the obvious thing to do and nothing
        here can see it happen, so the row outlives the file. Worse, the
        collector numbers the next run max(existing)+1 — delete the LAST
        episode and the next recording reuses its number, landing a second row
        with the same id. Surfacing both is the point of this.
        """
        return [r for r in self.read_rows()
                if r.get('status') != STATUS_DELETED
                and not os.path.isfile(os.path.join(self._dir, r['episode'] + '.hdf5'))]

    def reconcile(self) -> list:
        """Mark rows whose file is gone as deleted. Returns those episodes.

        The row is kept, not removed: 'we recorded this and threw it away' is
        a result, and a silently shrinking record is worse than an annotated
        one. Re-running is safe.
        """
        gone = [r['episode'] for r in self.orphan_rows()]
        if gone:
            rows = self.read_rows()
            for r in rows:
                if r['episode'] in gone:
                    r['status'] = STATUS_DELETED
            self._rewrite(rows)
            self._export_xlsx()
        return gone

    def append(self, row: dict):
        """Append one episode. Writes the csv first — that is the record."""
        row = dict(row)

        # An id can come round again after a hand-deleted episode. Keep the old
        # row as history but mark it, so exactly one row per episode is current
        # and a stale reading can never masquerade as the live one.
        episode = row.get('episode')
        prior = [r for r in self.read_rows()
                 if r['episode'] == episode and r.get('status') != STATUS_SUPERSEDED]
        if prior:
            rows = self.read_rows()
            for r in rows:
                if r['episode'] == episode:
                    r['status'] = STATUS_SUPERSEDED
            self._rewrite(rows)
            self._warn(
                f'{episode} already had {len(prior)} row(s) — the file was '
                f'probably deleted and the number reused. Older row(s) marked '
                f'{STATUS_SUPERSEDED}; the new one is current.')

        row.setdefault('recorded_at', datetime.now().isoformat(timespec='seconds'))
        row.setdefault('task', self._spec.name)
        row.setdefault('git_sha', self.git_sha)
        row.setdefault('git_dirty', self.git_dirty)

        os.makedirs(self._dir, exist_ok=True)
        is_new = not os.path.isfile(self._csv)
        try:
            with open(self._csv, 'a', newline='') as fh:
                w = csv.DictWriter(fh, fieldnames=self.columns, extrasaction='ignore')
                if is_new:
                    w.writeheader()
                w.writerow({c: row.get(c, '') for c in self.columns})
        except Exception as e:
            self._warn(f'Annotation NOT saved — could not write {self._csv}: {e}')
            return
        self._info(f"Annotated {row.get('episode')} → {self._csv}")
        self._export_xlsx()

    def _export_xlsx(self):
        """Regenerate the workbook from the csv. Best effort by design: the csv
        already holds the data, so every failure here is a warning, never a loss."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font
            from openpyxl.utils import get_column_letter
        except ImportError:
            if not self._xlsx_warned:
                self._xlsx_warned = True
                self._info(
                    'openpyxl not installed — annotations.csv is being written '
                    'normally; no .xlsx. To get one: /usr/bin/python3.12 -m pip '
                    'install --break-system-packages openpyxl')
            return

        try:
            with open(self._csv, newline='') as fh:
                rows = list(csv.reader(fh))
            if not rows:
                return

            wb = Workbook()
            ws = wb.active
            ws.title = os.path.basename(os.path.normpath(self._dir))[:31] or 'annotations'
            for r in rows:
                ws.append(r)

            bold = Font(bold=True)
            for cell in ws[1]:
                cell.font = bold
                cell.alignment = Alignment(horizontal='left')
            ws.freeze_panes = 'A2'
            ws.auto_filter.ref = ws.dimensions

            for i, name in enumerate(rows[0], start=1):
                widest = max((len(str(r[i - 1])) for r in rows if len(r) >= i),
                             default=len(name))
                ws.column_dimensions[get_column_letter(i)].width = min(
                    max(widest + 2, 10), 60)

            tmp = self._xlsx + '.tmp'
            wb.save(tmp)
            os.replace(tmp, self._xlsx)     # atomic: never a half-written workbook
        except Exception as e:
            self._warn(f'annotations.xlsx not updated ({e}) — annotations.csv is '
                       f'intact; close the workbook in Excel and it will refresh '
                       f'on the next episode.')


# ── Pygame prompt ──────────────────────────────────────────────────────────────

class AnnotationPrompt:
    """Modal panel drawn over the collector window after an episode is saved.

    Owns only its own state. Every field is answerable in one keystroke and in
    any order — there is no cursor to advance, because a cursor turns a
    two-second interaction into a five-second one and the operator is standing
    next to a live arm.
    """

    C_BG      = (28, 31, 36)
    C_INK     = (226, 228, 231)
    C_DIM     = (128, 134, 142)
    C_RULE    = (62, 68, 76)
    C_PICK    = (108, 168, 232)
    C_UNSET   = (232, 176, 72)
    C_OK      = (96, 200, 128)

    def __init__(self, spec: TaskSpec):
        self._spec = spec
        self.active = False
        self.meta = {}
        self.answers = {}
        self.notes = ''
        self._typing = False

    # -- lifecycle -------------------------------------------------------------
    def open(self, meta: dict):
        self.meta = dict(meta)
        self.answers = {}
        self.notes = ''
        self._typing = False
        self.active = True

    def close(self):
        self.active = False
        self._typing = False

    @property
    def capturing_text(self) -> bool:
        """True while the note is being typed. The caller must stop treating
        R / S / Q as escapes here, or typing 'restart' fires a robot reset."""
        return self._typing

    @property
    def unanswered(self):
        return [f for f in self._spec.fields if f.name not in self.answers]

    def row(self, status: str) -> dict:
        row = dict(self.meta)
        row['status'] = status
        row['notes'] = self.notes.strip()
        for f in self._spec.fields:
            row[f.name] = self.answers.get(f.name, '')
        return row

    # -- input -----------------------------------------------------------------
    def handle_key(self, event) -> str:
        """Returns 'save', 'skip', or '' (consumed / no decision yet).

        Returns '' for anything it handles itself. The caller keeps its own
        R/S/Q handling for everything else, so the prompt can never trap the
        operator away from the robot controls.
        """
        import pygame

        if self._typing:
            if event.key in (pygame.K_RETURN, pygame.K_KP_ENTER, pygame.K_ESCAPE):
                self._typing = False
            elif event.key == pygame.K_BACKSPACE:
                self.notes = self.notes[:-1]
            elif event.unicode and event.unicode.isprintable():
                self.notes = (self.notes + event.unicode)[:180]
            return ''

        if event.key in (pygame.K_RETURN, pygame.K_KP_ENTER):
            return 'save'
        if event.key == pygame.K_ESCAPE:
            return 'skip'

        ch = (event.unicode or '').lower()
        if ch == TEXT_KEY:
            self._typing = True
            return ''
        for f in self._spec.fields:
            value = f.value_for_key(ch)
            if value is not None:
                self.answers[f.name] = value
                return ''
        return ''

    # -- drawing ---------------------------------------------------------------
    def draw(self, screen, font, small, width, height):
        import pygame

        screen.fill(self.C_BG)
        pad = 20

        ep = self.meta.get('episode', '?')
        head = (f"{ep} saved  ·  {self.meta.get('num_frames', '?')} frames  ·  "
                f"{self.meta.get('duration_s', '?')} s")
        screen.blit(font.render(head, True, self.C_INK), (pad, 16))
        tag = small.render(self._spec.display, True, self.C_DIM)
        screen.blit(tag, (width - pad - tag.get_width(), 22))
        pygame.draw.line(screen, self.C_RULE, (pad, 46), (width - pad, 46))

        y = 62
        for f in self._spec.fields:
            screen.blit(small.render(f.ask, True, self.C_INK), (pad, y))
            y += 22
            x = pad + 8
            for value, key in f.choices:
                picked = self.answers.get(f.name) == value
                label = f'[{key.upper()}] {value}'
                surf = small.render(label, True,
                                    self.C_PICK if picked else self.C_DIM)
                if picked:
                    pygame.draw.rect(screen, self.C_PICK,
                                     (x - 5, y - 3, surf.get_width() + 10,
                                      surf.get_height() + 6), 1)
                screen.blit(surf, (x, y))
                x += surf.get_width() + 26
            if f.rule:
                y += 22
                screen.blit(small.render(f.rule, True, self.C_RULE), (pad + 8, y))
            y += 30

        # Notes
        screen.blit(small.render('Anything odd this run?', True, self.C_INK), (pad, y))
        y += 22
        box = pygame.Rect(pad + 8, y - 4, width - 2 * pad - 16, 26)
        pygame.draw.rect(screen, self.C_PICK if self._typing else self.C_RULE, box, 1)
        shown = self.notes if (self.notes or self._typing) else '[T] to type'
        colour = self.C_INK if self.notes else self.C_RULE
        if self._typing:
            shown += '_'
        screen.blit(small.render(shown[-72:], True, colour), (box.x + 8, box.y + 4))
        y += 40

        # Footer — what pressing things will do, and what will be written.
        pygame.draw.line(screen, self.C_RULE, (pad, height - 54),
                         (width - pad, height - 54))
        missing = self.unanswered
        if self._typing:
            state, colour = 'typing — ENTER to finish the note', self.C_PICK
        elif missing:
            state = 'unanswered: ' + ', '.join(f.name for f in missing)
            colour = self.C_UNSET
        else:
            state, colour = 'all answered', self.C_OK
        screen.blit(small.render(state, True, colour), (pad, height - 46))
        screen.blit(small.render(
            'ENTER save   ·   ESC skip (marks unreviewed)   ·   R / S / Q also exit',
            True, self.C_DIM), (pad, height - 26))


# ── CLI ────────────────────────────────────────────────────────────────────────

def main():
    import argparse
    ap = argparse.ArgumentParser(
        description='Inspect or reconcile a collection\'s annotations.')
    ap.add_argument('collection', help='folder holding episode_*.hdf5')
    ap.add_argument('--task', default='grape_pluck', help='task spec name')
    ap.add_argument('--reconcile', action='store_true',
                    help='mark rows whose episode file has been deleted')
    ap.add_argument('--export', action='store_true',
                    help='regenerate annotations.xlsx from the csv')
    args = ap.parse_args()

    store = AnnotationStore(args.collection, TaskSpec.load(args.task))
    rows = store.read_rows()
    if not rows:
        print(f'No annotations.csv in {args.collection}')
        return

    if args.reconcile:
        gone = store.reconcile()
        print(f'Marked {len(gone)} row(s) {STATUS_DELETED}: {", ".join(gone)}'
              if gone else 'Nothing to reconcile — every row has its file.')
    elif args.export:
        store._export_xlsx()
        print(f'Wrote {os.path.join(args.collection, "annotations.xlsx")}')

    rows = store.read_rows()
    print(f'\n{len(rows)} row(s) in {args.collection}')
    for r in rows:
        on_disk = os.path.isfile(
            os.path.join(args.collection, r['episode'] + '.hdf5'))
        flag = '' if on_disk else '   [no file]'
        print(f"  {r['episode']:12s} {r.get('status',''):11s} "
              f"{r.get('num_frames',''):>5s}f  {r.get('rate_hz',''):>5s}Hz  "
              f"{r.get('notes','')[:32]}{flag}")
    orphans = store.orphan_rows()
    if orphans:
        print(f'\n{len(orphans)} row(s) reference a missing episode. '
              f'Run with --reconcile to mark them.')


if __name__ == '__main__':
    main()
